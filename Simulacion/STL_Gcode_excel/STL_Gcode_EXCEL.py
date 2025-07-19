import subprocess
import os
import re # Para expresiones regulares
import openpyxl # Para manejar archivos Excel (.xlsx)
from openpyxl.styles import Font, Alignment

# --- Modificación 1: analyze_gcode_and_report_extended ahora toma la ruta completa del G-code ---
def analyze_gcode_and_report_extended(gcode_full_path):
    """
    Extrae tiempo estimado (normal y silencioso), y filamento usado (mm, cm3, g)
    de un archivo .gcode y exporta los datos a un archivo Excel.
    Acepta la ruta completa del archivo G-code como argumento.
    El archivo Excel se guardará en el mismo directorio que el G-code,
    con el mismo nombre base que el G-code pero con extensión .xlsx.
    """
    
    # --- Determinar rutas de salida para el reporte ---
    gcode_dir = os.path.dirname(gcode_full_path) # Directorio donde se encuentra el G-code
    gcode_filename_base = os.path.splitext(os.path.basename(gcode_full_path))[0] # Nombre base sin extensión
    
    excel_filename = f"{gcode_filename_base}.xlsx" # Nombre del archivo Excel
    excel_path = os.path.join(gcode_dir, excel_filename) # Ruta completa del archivo Excel

    # El nombre que se mostrará en Excel será el nombre base del G-code
    order_code_name = gcode_filename_base

    print(f"\nAnalizando archivo G-code: {os.path.basename(gcode_full_path)}")

    # --- Extraer datos del G-code ---
    # Inicializar todos los valores a "N/A" (No Aplicable/No Encontrado)
    filament_mm = "N/A"
    filament_cm3 = "N/A"
    filament_g = "N/A"
    estimated_time_normal = "N/A"
    estimated_time_silent = "N/A"

    try:
        with open(gcode_full_path, 'r', encoding='utf-8') as f:
            for line in f:
                # Buscar filamento usado [mm]
                mm_match = re.search(r'; filament used \[mm\] = (\d+\.?\d*)', line)
                if mm_match:
                    filament_mm = mm_match.group(1).strip()

                # Buscar filamento usado [cm3]
                cm3_match = re.search(r'; filament used \[cm3\] = (\d+\.?\d*)', line)
                if cm3_match:
                    filament_cm3 = cm3_match.group(1).strip()

                # Buscar filamento usado [g]
                g_match = re.search(r'; filament used \[g\] = (\d+\.?\d*)?', line)
                if g_match:
                    filament_g = g_match.group(1).strip() if g_match.group(1) else "0" # Si está vacío, asignar "0"

                # Buscar tiempo estimado (normal mode)
                normal_time_match = re.search(r'; estimated printing time \(normal mode\) = (.+)', line)
                if normal_time_match:
                    estimated_time_normal = normal_time_match.group(1).strip()

                # Buscar tiempo estimado (silent mode)
                silent_time_match = re.search(r'; estimated printing time \(silent mode\) = (.+)', line)
                if silent_time_match:
                    estimated_time_silent = silent_time_match.group(1).strip()
                
                # Opcional: Detener la lectura si se encuentran todos los datos clave
                # Solo detener si todos son diferentes de "N/A"
                if (filament_mm != "N/A" and filament_cm3 != "N/A" and filament_g != "N/A" and
                    estimated_time_normal != "N/A" and estimated_time_silent != "N/A"):
                    break

    except Exception as e:
        print(f"Error al leer el archivo G-code: {e}")
        return

    print(f"  Filamento usado [mm]: {filament_mm}")
    print(f"  Filamento usado [cm3]: {filament_cm3}")
    print(f"  Filamento usado [g]: {filament_g}")
    print(f"  Tiempo estimado (normal mode): {estimated_time_normal}")
    print(f"  Tiempo estimado (silent mode): {estimated_time_silent}")

    # --- Exportar a Excel ---
    try:
        # Asegúrate de que el directorio del G-code exista (ya debería existir)
        os.makedirs(gcode_dir, exist_ok=True)

        # Crear o cargar un libro de trabajo de Excel
        if os.path.exists(excel_path):
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active # Obtener la hoja activa
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reporte" # Establecer el título de la hoja

            # Escribir encabezados si es un archivo nuevo
            headers = [
                "Codigo de orden",
                "Filamento usado [mm]",
                "Filamento usado [cm3]",
                "Filamento usado [g]",
                "Tiempo estimado (normal)",
                "Tiempo estimado (silencioso)"
            ]
            
            # Escribir encabezados en la primera fila
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas para encabezados
            column_widths = {
                'A': 30, 'B': 25, 'C': 25, 'D': 25, 'E': 30, 'F': 30
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width


        # Añadir los nuevos datos
        next_row = sheet.max_row + 1 # La siguiente fila disponible
        sheet.cell(row=next_row, column=1, value=order_code_name) # Usa el nombre base del archivo para el "Codigo de orden"
        sheet.cell(row=next_row, column=2, value=filament_mm)
        sheet.cell(row=next_row, column=3, value=filament_cm3)
        sheet.cell(row=next_row, column=4, value=filament_g)
        sheet.cell(row=next_row, column=5, value=estimated_time_normal)
        sheet.cell(row=next_row, column=6, value=estimated_time_silent)

        # Guardar el libro de trabajo
        workbook.save(excel_path)
        print(f"\nDatos exportados exitosamente a: {excel_path}")

    except Exception as e:
        print(f"Error al exportar a Excel: {e}")

# --- Modificación: slice_model_and_report ahora acepta order_code como parámetro ---
def slice_model_and_report(order_code):
    """
    'Slicea' un modelo STL basado en el código de orden y luego genera un reporte Excel.
    El STL se busca en 'pedidos/order_code/order_code.stl'.
    El G-code y el Excel se guardan en 'pedidos/order_code/'.
    """
    # Obtener la ruta absoluta del directorio donde se encuentra este script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    print(f"Directorio del script: {script_dir}\n")

    # --- Rutas relativas y absolutas ---
    prusa_slicer_exe_relative = os.path.join("PrusaSlicer-2.9.1", "prusa-slicer-console.exe")
    config_file_relative = "config.ini"
    
    pedidos_base_dir_relative = "pedidos" 
    
    # La carpeta específica para este pedido (ej. 'pedidos/ORDEN123/')
    order_specific_dir_relative = os.path.join(pedidos_base_dir_relative, order_code)

    # Construir las rutas absolutas
    prusa_slicer_exe = os.path.join(script_dir, prusa_slicer_exe_relative)
    config_file = os.path.join(script_dir, config_file_relative)
    
    # Ruta completa de la carpeta específica del pedido
    order_specific_dir = os.path.join(script_dir, order_specific_dir_relative)

    # Ruta completa del archivo STL de entrada: pedidos/order_code/order_code.stl
    input_stl_filename = f"{order_code}.stl"
    input_stl = os.path.join(order_specific_dir, input_stl_filename)
    
    # Generar el nombre del archivo G-code dinámicamente
    output_gcode_filename = f"{order_code}.gcode"
    
    # Ruta completa del archivo G-code de salida (dentro de la carpeta del pedido)
    output_gcode_path = os.path.join(order_specific_dir, output_gcode_filename)
    
    # Asegúrate de que el directorio específico del pedido exista
    os.makedirs(order_specific_dir, exist_ok=True)

    # Verificar si el archivo STL de entrada existe antes de ejecutar PrusaSlicer
    if not os.path.exists(input_stl):
        print(f"Error: El archivo STL '{input_stl}' no se encontró.")
        print(f"Asegúrate de que el archivo '{input_stl_filename}' esté en la carpeta '{order_specific_dir_relative}' y que el nombre de la orden sea correcto.")
        return False # Indicar que la operación falló

    # Construir la lista de argumentos para el comando
    command = [
        prusa_slicer_exe,
        "--load", config_file,
        "--export-gcode", input_stl,
        "--output", output_gcode_path
    ]

    print(f"\nComando de PrusaSlicer a ejecutar: {' '.join(command)}\n")

    try:
        # Ejecutar el comando PrusaSlicer
        result = subprocess.run(command, capture_output=True, text=True, encoding='utf-8')

        print("--- Salida estándar del comando de PrusaSlicer ---")
        print(result.stdout)
        print("--- Errores estándar del comando de PrusaSlicer ---")
        print(result.stderr)
        
        if result.returncode != 0:
            print(f"Error: El comando de PrusaSlicer terminó con código de retorno {result.returncode}.")
            print("No se pudo generar el G-code. Por favor, revisa los errores anteriores.")
            return False # Indicar que la operación falló

        print("\nComando de PrusaSlicer ejecutado exitosamente.")
        print(f"Archivo G-code generado: {output_gcode_path}")

        # Llamar a la función de reporte después de generar el G-code
        analyze_gcode_and_report_extended(output_gcode_path)
        return True # Indicar que la operación fue exitosa

    except FileNotFoundError:
        print(f"Error: El ejecutable de PrusaSlicer no se encontró en la ruta: {prusa_slicer_exe}")
        print("Verifica que la ruta sea correcta y que la carpeta 'PrusaSlicer-2.9.1' esté en el mismo nivel que tu script.")
        return False # Indicar que la operación falló
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
        return False # Indicar que la operación falló

# Eliminar la llamada directa a slice_model_and_report() para que sea llamada desde el script principal
# if __name__ == "__main__":
#     slice_model_and_report()
